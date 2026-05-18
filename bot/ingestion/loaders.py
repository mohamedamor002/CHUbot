"""
Chargement de documents RH — fichiers locaux et pages web.

Formats supportés : PDF, DOCX, DOC, XLSX, XLS, XLSM, pages web statiques/JS.
"""

import time
from pathlib import Path
from typing import List
from urllib.parse import urljoin, urlparse

import requests
import trafilatura
from bs4 import BeautifulSoup
from langchain_core.documents import Document
from langchain_community.document_loaders import PyMuPDFLoader, Docx2txtLoader


SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".doc", ".xlsx", ".xls", ".xlsm"}

_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}


# ── Fichiers locaux ───────────────────────────────────────────────────────────

def _load_excel(path: Path) -> List[Document]:
    """Charge un fichier Excel avec openpyxl — plus fiable qu'UnstructuredExcelLoader."""
    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    docs = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = []
        for row in ws.iter_rows(values_only=True):
            row_text = "\t".join(str(c) if c is not None else "" for c in row).strip()
            if row_text:
                rows.append(row_text)
        if rows:
            docs.append(Document(
                page_content="\n".join(rows),
                metadata={"source_file": path.name, "file_type": path.suffix.lstrip("."), "sheet": sheet},
            ))
    wb.close()
    return docs


def load_file(file_path: str) -> List[Document]:
    """Charge un fichier (PDF, DOCX, Excel) et retourne des Documents LangChain."""
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Fichier introuvable : {file_path}")

    ext = path.suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"Format non supporté : {ext}. Acceptés : {SUPPORTED_EXTENSIONS}")

    if ext == ".pdf":
        loader = PyMuPDFLoader(str(path))
    elif ext in {".docx", ".doc"}:
        loader = Docx2txtLoader(str(path))
    else:
        return _load_excel(path)

    docs = loader.load()
    for doc in docs:
        doc.metadata["source_file"] = path.name
        doc.metadata["file_type"] = ext.lstrip(".")
    return docs


def load_directory(dir_path: str) -> List[Document]:
    """Charge tous les fichiers supportés d'un dossier (récursif)."""
    directory = Path(dir_path)
    if not directory.is_dir():
        raise NotADirectoryError(f"Dossier introuvable : {dir_path}")

    all_docs: List[Document] = []
    for file in directory.rglob("*"):
        if file.suffix.lower() in SUPPORTED_EXTENSIONS:
            try:
                docs = load_file(str(file))
                all_docs.extend(docs)
                print(f"  [OK] {file.name} ({len(docs)} page(s))")
            except Exception as e:
                print(f"  [ERR] {file.name} — {e}")
    return all_docs


# ── Pages web statiques (trafilatura) ────────────────────────────────────────

def _scrape_static(url: str) -> str | None:
    response = requests.get(url, headers=_HEADERS, timeout=15)
    response.raise_for_status()
    return trafilatura.extract(
        response.text,
        include_links=True,
        include_tables=True,
        include_images=False,
        deduplicate=True,
    )


# ── Pages web JavaScript (Playwright) ────────────────────────────────────────

def _scrape_js(url: str, wait_seconds: int = 3) -> str | None:
    from playwright.sync_api import sync_playwright

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        page.goto(url, wait_until="networkidle", timeout=30000)
        time.sleep(wait_seconds)
        html = page.content()
        browser.close()

    return trafilatura.extract(html, include_links=True, include_tables=True, deduplicate=True)


def load_url(url: str, use_js: bool = False) -> List[Document]:
    """Scrape une URL et retourne des Documents LangChain."""
    try:
        text = _scrape_js(url) if use_js else _scrape_static(url)
    except Exception as e:
        raise RuntimeError(f"Erreur scraping {url}: {e}") from e

    if not text or len(text.strip()) < 100:
        return []

    try:
        r = requests.get(url, headers=_HEADERS, timeout=10)
        soup = BeautifulSoup(r.text, "lxml")
        title = soup.title.get_text(strip=True) if soup.title else urlparse(url).path
    except Exception:
        title = urlparse(url).path

    return [Document(
        page_content=text,
        metadata={"source_file": title, "source_url": url, "file_type": "web"},
    )]


def load_urls(urls: List[str], use_js: bool = False) -> List[Document]:
    """Scrape une liste d'URLs spécifiques."""
    all_docs: List[Document] = []
    for url in urls:
        try:
            docs = load_url(url, use_js=use_js)
            all_docs.extend(docs)
            print(f"  [OK] {url} ({len(docs)} doc(s))")
        except Exception as e:
            print(f"  [ERR] {url} — {e}")
    return all_docs


def load_site(base_url: str, max_pages: int = 30, use_js: bool = False) -> List[Document]:
    """Crawle un site entier à partir d'une URL de base."""
    base_domain = urlparse(base_url).netloc
    visited: set = set()
    to_visit = [base_url]
    found: List[str] = []

    print(f"  Découverte des sous-pages de {base_url}...")
    while to_visit and len(found) < max_pages:
        url = to_visit.pop(0)
        if url in visited:
            continue
        visited.add(url)
        try:
            r = requests.get(url, headers=_HEADERS, timeout=10)
            if "text/html" not in r.headers.get("Content-Type", ""):
                continue
            found.append(url)
            soup = BeautifulSoup(r.text, "lxml")
            for a in soup.find_all("a", href=True):
                href = urljoin(url, a["href"]).split("#")[0].split("?")[0]
                if (
                    urlparse(href).netloc == base_domain
                    and href not in visited
                    and href not in to_visit
                    and not href.endswith((".pdf", ".docx", ".xlsx", ".jpg", ".png"))
                ):
                    to_visit.append(href)
        except Exception:
            continue

    print(f"  {len(found)} pages trouvées")
    return load_urls(found, use_js=use_js)


def download_linked_files(
    page_url: str,
    save_dir: str | Path,
    extensions: tuple = (".docx", ".doc", ".pdf", ".xlsx", ".xlsm", ".xls"),
) -> List[Path]:
    """Télécharge les fichiers liés (.pdf, .docx…) trouvés sur une page."""
    save_path = Path(save_dir)
    save_path.mkdir(parents=True, exist_ok=True)

    try:
        r = requests.get(page_url, headers=_HEADERS, timeout=15)
        r.raise_for_status()
    except Exception as e:
        print(f"  [ERR] Accès impossible à {page_url}: {e}")
        return []

    soup = BeautifulSoup(r.text, "lxml")
    downloaded: List[Path] = []

    for a in soup.find_all("a", href=True):
        href = a["href"].split("?")[0].split("#")[0]
        if not any(href.lower().endswith(ext) for ext in extensions):
            continue
        file_url = urljoin(page_url, href)
        filename = Path(urlparse(file_url).path).name
        if not filename:
            continue
        dest = save_path / filename
        if dest.exists():
            print(f"  [SKIP] {filename}")
            downloaded.append(dest)
            continue
        try:
            resp = requests.get(file_url, headers=_HEADERS, timeout=30)
            resp.raise_for_status()
            dest.write_bytes(resp.content)
            print(f"  [DL] {filename} ({len(resp.content) // 1024} Ko)")
            downloaded.append(dest)
        except Exception as e:
            print(f"  [ERR] {filename} — {e}")

    return downloaded
